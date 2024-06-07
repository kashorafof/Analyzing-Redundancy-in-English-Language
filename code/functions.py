import os
from os.path import exists
import csv
from mainConfig import *
import spacy 
import openpyxl
import matplotlib.pyplot as plt
import time
from scipy.stats import sem
import numpy as np
import json
from spacy.tokens import Doc

# used to write a dictionary to a file
def save(File_name, dict):
    with open(File_name , 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(dict.keys())
        writer.writerow(dict.values())

# used to load a dictionary from a file
def load(File_name):
    # in case no file -> create file with empty dictionary and load it 
    if not exists(File_name ):
        save(File_name, dict())
        return dict()

    with open(File_name , 'r') as csvfile:
        reader = csv.reader(csvfile)
        return dict(zip(next(reader), next(reader)))

# used to write a list to a json file
def docSave(docs, file):
    with open(file, 'w', newline='', errors = 'replace') as jsonFile:
        json.dump([doc.to_json() for doc in docs], jsonFile)
        
# used to load a list from a json file
def docLoad(file):
    with open(file, 'r', newline='', errors = 'replace') as jsonFile:
        docs_json = json.load(jsonFile)
        return [ sent for doc_json in docs_json for sent in Doc(nlp.vocab).from_json(doc_json).sents]



# to filter the links and remove unnecessary links "png, jpg ,.. etc"
def filter_links():
    for website_name in website_list.keys():
        path = result_path + '/scrapped/' +website_name + '/' + website_name + '_Links.txt'
        file = open(path, 'r+' , encoding='utf-8' )
        s = ''
        text = file.readlines()
        for line in text:
            line = line.strip()
            # check if the link is not a picture or a video .. etc
            if line.lower().endswith(forbidden_ends):
                continue
            s += line + '\n'
        open( path , 'w+', encoding= 'utf-8').write(s)


# count the number of articles in each category for specific website
def count_category(website, category):
    path = result_path + '/scrapped/'+website + '/texts/' + category + '/'
    if not exists(path):
        return 0
    return len(os.listdir(path))

# count the number of links for all websites
# return a dictionary with the number of links for each website
def num_link():
    num_link = dict.fromkeys(website_list.keys(), 0)
    for website_name in website_list.keys():
        path = result_path + '/scrapped/'+website_name + '/' + website_name + '_Links.txt'
        file = open(path, 'r+' , encoding='utf-8' )
        text = file.readlines()
        num_link[website_name] = len(text)
    return num_link


# count the number of articles for all websites
# return a dictionary with the number of articles for each category
def num_Articles():
    categories_Dict = dict.fromkeys(categories, 0)
    for category in categories:
        for webste_name in website_list.keys():
            path = result_path + '/scrapped/' + webste_name + '/texts/' + category + '/'
            if not exists(path):
                continue
            categories_Dict[category] += len(os.listdir(path))
    return categories_Dict

# combine all the articles from a specific category and return it as list
def combine_categ(categ):
    txts = [open(result_path + '/scrapped/'+ webSite + '/texts/' + categ + '/' + filename, 'r+', encoding='utf-8', errors='replace').read().strip().lower() for webSite in website_list.keys() for filename in os.listdir(result_path + '/' + webSite + '/texts/' + categ + '/') if filename.endswith(".txt")]
    return txts



# used to get the data from the abbreviation excel file and return a dictionary with the abbreviations
def get_data(location):

    abbreviations = {}
    workbook = openpyxl.load_workbook(location)
    sheet = workbook.active

    for i in range(2,sheet.max_row):
        abbreviations[sheet.cell(i,1).value.strip().lower()] = sheet.cell(i,2).value.strip().lower()
    
    return abbreviations

# 
def analyse_txts(txts):
    # pass the text through nlp pipe to get the sentences after the processing
    piped = [pipe for pipe in nlp.pipe(txts,n_process = number_threads, batch_size= 20)]
    # return the senteces of the text after the processing
    return [sent for pipe in piped for sent in pipe.sents]

# used to get the statistics of the words in the articles of each category
# used as the main driver for getting the statistics
def get_word_statistics():
    print ('start1')
    t = time.time()
    start = time.time()
    # load the abbreviation dictionary
    abbrv = get_data(abbreviation_location)
    
    # run the analysis for each category
    
    for category in categories:
        t = time.time() ###
        # get all the articles from the category
        txts = combine_categ(category)
        # get the sentences of the articles after processing
        docs = analyse_txts(txts)

        # combine all the articles into one text
        # will use it to get the number of words in the text
        txt = '\n'.join(txts)
        txtLength = len(txt.split())
        
        print( category + ' analys done in: ' + str(time.time() - t)) ###
        t = time.time() ###
        # start the analysis for each rule 
        for rule in rules_results.keys():

            # call the function of the rule -check file text_processing.py-
            # rules_fun is a dictionary that map each rule to its function
            if rule == 'Abbreviation':
                rule_occ = rules_fun[rule] (txt, abbrv) 
            else:
                rule_occ = sum( rules_fun[rule] (doc) for doc in docs)

            # calculate the number of occurrences of the rule in the text per 100 word 
            rules_results[rule][category] = 100 * rule_occ / txtLength

            print( category + '\t' + str(rules_abbrv[rule]) + '\t' + str(time.time() - t) + '\tocc: ' + str(rules_results[rule][category]) + '\n') ###
        
    return rules_results